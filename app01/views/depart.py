from django.shortcuts import render, redirect

from openpyxl import load_workbook
from app01.models import Depart
from app01.utils.pagination import Pagination


def depart_list(request):
    depart_data = Depart.objects.all()

    page_obj = Pagination(request, depart_data)

    context = {
        'depart_data': page_obj.page_queryset,
        'page_string': page_obj.html()
    }

    return render(request, 'depart_list.html', context)


def depart_add(request):
    if request.method == "GET":
        return render(request, 'depart_add.html')

    depart_title = request.POST.get('title')
    exists = Depart.objects.filter(title=depart_title).exists()
    if not exists:
        Depart.objects.create(title=depart_title)
    else:
        return render(request, 'depart_add.html', {'error_msg': '部门名称存在请重新输入'})

    return redirect('/depart/list/')


def depart_delete(request, nid):
    Depart.objects.filter(id=nid).delete()
    return redirect('/depart/list/')


def depart_edit(request, nid):
    if request.method == "GET":
        depart_info = Depart.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html', {'depart_info': depart_info})

    title_new = request.POST.get('title')
    Depart.objects.filter(id=nid).update(title=title_new)
    return redirect('/depart/list/')


def depart_multi(request):
    file_obj = request.FILES.get('file_content')

    if file_obj:

        wb = load_workbook(file_obj)
        sheet = wb.worksheets[0]

        for row in sheet.iter_rows(min_row=2):
            text = row[0].value
            print(text)
            exists = Depart.objects.filter(title=text).exists()
            if not exists:
                Depart.objects.create(title=text)

        return redirect('/depart/list/')

    return redirect('/depart/list/')

